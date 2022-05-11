'''
This program uses selenium webdriver to:
1. load and login to metamask browser extension
2. connect to alphahomora v2 on AVAX C-Chain
3. determine the current APR offered for each LP (liquidity provider) pair
4. update an excel sheet with the current APRs

Bugs may occur if:
1. exchanges are added/removed/renamed
2. LP pairs are added/removed/renamed

Adjust sleep times if depending on how fast selenium webdriver loads.

This program only works if you already have an EVM compatible wallet
secret phrase.
'''

from selenium import webdriver
from time import sleep
import json
import pandas as pd
from datetime import date
import openpyxl

# load config
f = open('metamask_config.json')
metamask_config = json.load(f)
f.close()

# paths
EXTENSION_PATH = 'E:\\pythonFiles\\metamask-10.9.3-an+fx.xpi'
alpha = 'https://homora-v2.alphafinance.io/dashboard'


def start():
    # start browser with metamask
    driver = webdriver.Firefox()
    driver.install_addon(EXTENSION_PATH, temporary=True)
    sleep(10)

    # switch to metamask initialization window
    metamask_window = driver.window_handles[1]
    driver.switch_to.window(metamask_window)

    # setup metamask
    driver.find_element_by_xpath('//button[text()="Get Started"]').click()
    driver.find_element_by_xpath('//button[text()="Import wallet"]').click()
    driver.find_element_by_xpath('//button[text()="No Thanks"]').click()
    inputs = driver.find_elements_by_xpath('//input')
    inputs[0].send_keys(metamask_config['secret'])
    inputs[1].send_keys('password')
    inputs[2].send_keys('password')
    driver.find_element_by_css_selector('.first-time-flow__terms').click()
    driver.find_element_by_xpath('//button[text()="Import"]').click()
    sleep(10)
    driver.find_element_by_xpath('//button[text()="All Done"]').click()

    return driver


def connect(driver):
    # open alpha homora
    driver.get(alpha)
    sleep(8)
    popup_window = driver.window_handles[2]
    driver.switch_to.window(popup_window)

    # connect to alpha homora ETH
    driver.find_element_by_xpath('//button[text()="Next"]').click()
    driver.find_element_by_xpath('//button[text()="Connect"]').click()
    main_window = driver.window_handles[1]
    driver.switch_to.window(main_window)
    sleep(10)

    # switch to AVAX
    driver.find_element_by_xpath('//span[text()="Ethereum"]').click()
    driver.find_element_by_xpath('//p[text()="Avalanche"]').click()
    sleep(15)
    popup_window = driver.window_handles[2]
    driver.switch_to.window(popup_window)
    driver.find_element_by_xpath('//button[text()="Approve"]').click()
    driver.find_element_by_xpath('//button[text()="Switch network"]').click()
    sleep(20)

    # switch back to main window for data collection
    main_window = driver.window_handles[1]
    driver.switch_to.window(main_window)


def collectData(driver):
    apy = []
    names = []
    pair = {}

    # find net APYs
    for i in driver.find_elements_by_tag_name('h4'):
        if i.text[-1] == '%':
            apy.append(i.text[:-2])

    # find names of LPs
    for i in driver.find_elements_by_tag_name('p'):
        if 'Trader Joe' in i.text or 'Pangolin V2' in i.text:
            names.append(i.text)

    # creates a dictionary pairing appropriate LP names and APYs
    # use apy to avoid using active position LP
    for i in range(len(apy)):
        pair[names[i]] = float(apy[i])
    return pair


def saveData(pair):
    # set up dataframe where today's date is index
    today = date.today().strftime("%m/%d/%y")
    pair['Date'] = today
    df = pd.DataFrame(data=pair, index=[0])
    df = df.set_index('Date')
    # quick check to make sure data is saved properly
    print(df)
    filename = 'Alpha Homora APR.xlsx'
    sheetname = 'Historical APR'
    # write to excel file, tries to add to sheet, makes new file if file doesn't exist
    try:
        wb = openpyxl.load_workbook(filename=filename)
        sheet = wb[sheetname]
        new_row = df.reset_index().iloc[0].tolist()
        # checks to see if there are more LPs scraped than saved previously
        if len(new_row) != sheet.max_column:
            with pd.ExcelWriter(filename, mode='a') as writer:
                df.to_excel(writer, sheet_name=sheetname)
            df.to_excel(filename, sheet_name=f'{sheetname}{len(wb.worksheets)}')  # makes new sheet if there is
        # add new row to top of list, not removing average column
        else:
            sheet.insert_rows(3)
            sheet.append(new_row)
            alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M',
                        'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
            sheet.move_range(f'A{sheet.max_row}:{alphabet[len(new_row)]}{sheet.max_row}', rows=(sheet.min_row - sheet.max_row + 2))
            wb.save(filename)

    except FileNotFoundError:
        df.to_excel(filename, sheet_name=sheetname)


def main():
    driver = start()
    connect(driver)

    pair = collectData(driver)
    print(pair)
    saveData(pair)

    driver.quit()


main()
